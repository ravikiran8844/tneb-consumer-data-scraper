import asyncio
from playwright.async_api import async_playwright, TimeoutError as PlaywrightTimeoutError
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import os

EXCEL_FILE = "consumer_data.xlsx"

async def scrape():
    # Create or load Excel file and worksheet
    if os.path.exists(EXCEL_FILE):
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        # Write headers
        ws.append([
            "Consumer No",
            "Consumer Name",
            "Address",
            "Sanctioned Load",
            "Tariff Details",
            "Service Effected Date"
        ])
        wb.save(EXCEL_FILE)

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=False)
        page = await browser.new_page()

        url = "https://www.tnebltd.gov.in/HTdetails/htconsumerindex.xhtml"

        for i in range(1, 10000):
            consumer_no = str(i).zfill(4)
            print(f"Fetching data for: {consumer_no}")

            try:
                await page.goto(url, wait_until="networkidle")

                # Click district label and option
                await page.click('#j_idt5\\:district_label')
                await page.wait_for_selector('#j_idt5\\:district_1')
                await page.click('#j_idt5\\:district_1')

                # Click section label and option
                await page.click('#j_idt5\\:section_label')
                await page.wait_for_selector('#j_idt5\\:section_1')
                await page.click('#j_idt5\\:section_1')

                # Set consumer number input value
                await page.fill('#j_idt5\\:scno', consumer_no)

                # Submit the form
                await page.click('#j_idt5\\:submit1')

                # Wait for results table to load (timeout 10 sec)
                await page.wait_for_selector('#j_idt6\\:panel_content > table', timeout=10000)

                # Extract data using page.eval_on_selector (JS in the browser context)
                consumer_name = await page.eval_on_selector(
                    '#j_idt6\\:panel_content > table > tbody > tr:nth-child(1) > td:nth-child(2) > label',
                    'el => el.textContent.trim()'
                )
                address = await page.eval_on_selector(
                    '#j_idt6\\:panel_content > table > tbody > tr:nth-child(2) > td:nth-child(2) > label',
                    'el => el.textContent.trim()'
                )
                sanctioned_load = await page.eval_on_selector(
                    '#j_idt6\\:panel_content > table > tbody > tr:nth-child(3) > td:nth-child(2) > label',
                    'el => el.textContent.trim()'
                )
                tariff_details = await page.eval_on_selector(
                    '#j_idt6\\:panel_content > table > tbody > tr:nth-child(4) > td:nth-child(2) > label',
                    'el => el.textContent.trim()'
                )
                # Service Effected Date selector fallback span or label
                service_effected_date = await page.eval_on_selector(
                    '#j_idt6\\:panel_content > table > tbody > tr:nth-child(5) > td:nth-child(2) > span',
                    'el => el.textContent.trim()',
                    strict=False
                )
                if not service_effected_date:
                    service_effected_date = await page.eval_on_selector(
                        '#j_idt6\\:panel_content > table > tbody > tr:nth-child(5) > td:nth-child(2) > label',
                        'el => el.textContent.trim()',
                        strict=False
                    )
                # Save data to Excel
                ws.append([
                    consumer_no,
                    consumer_name,
                    address,
                    sanctioned_load,
                    tariff_details,
                    service_effected_date
                ])
                wb.save(EXCEL_FILE)
                print(f"Saved data for {consumer_no}")

            except PlaywrightTimeoutError:
                print(f"Timeout or no data for consumer {consumer_no}, skipping.")
            except Exception as e:
                print(f"Error for consumer {consumer_no}: {e}")

        await browser.close()

if __name__ == "__main__":
    asyncio.run(scrape())
